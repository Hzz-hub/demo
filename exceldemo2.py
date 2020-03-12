import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
print(ws.title)
print(ws.values)
ws['a1'] = 111
print(ws['a1'].value)
for row in range(1, 10):
    ws.cell(row, 1).value = row
    print(ws.cell(row, 1).value)

wb.save('transaction.xlsx')
