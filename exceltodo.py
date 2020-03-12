import openpyxl as xl


wb = xl.load_workbook('transaction.xlsx')
wb.create_sheet(title='总结', index=0)
ws = wb.sheetnames
print(ws)
wb.remove(wb['总结'])
ws = wb.sheetnames
print(ws)